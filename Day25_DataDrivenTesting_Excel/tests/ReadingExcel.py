import openpyxl

# File--->WorkBook--->Sheets--->Rows--->Cells

file = r'D:\DELL DATA\Documents\PavanKumar_Python\Day25-DataDrivenTesting_Excel\ClassExamples (2)\ClassExamples\data.xlsx'

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

# Count the number of rows in an Excel sheet 6
rows = sheet.max_row
print(rows)

# Count the number columns in an Excel sheet 4
cols = sheet.max_column

# Reading all the rows and columns from excel sheet
for r in range(1, rows + 1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end="      ")
    print("")
