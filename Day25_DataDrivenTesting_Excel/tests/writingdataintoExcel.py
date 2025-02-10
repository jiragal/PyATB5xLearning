import openpyxl
from openpyxl.styles import PatternFill

# File--->WorkBook--->Sheets--->Rows--->Cells

file1 = r'D:\DELL DATA\Documents\PavanKumar_Python\Day25-DataDrivenTesting_Excel\ClassExamples (2)\ClassExamples\testdata1.xlsx'
file = r'D:\DELL DATA\Documents\PavanKumar_Python\Day25-DataDrivenTesting_Excel\ClassExamples (2)\ClassExamples\data.xlsx'

workbook1 = openpyxl.load_workbook(file)
sheet = workbook1["Sheet3"]  # if only one sheet we can read like this "workbook1.active"
for r in range(1, 6):
    for c in range(1, 5):
        sheet.cell(r, c).value = "IBM"

workbook1.save(file)

workbook2 = openpyxl.load_workbook(file1)

sheet1 = workbook2.active  # (or) sheet=workbook["Sheet4"]     --  get active sheet from excel

sheet1.cell(1, 1).value = 'NAME'
sheet1.cell(1, 2).value = 'COMPANY'
sheet1.cell(1, 3).value = 'EXPERIENCE'

sheet1.cell(2, 1).value = 'NINGU'
sheet1.cell(2, 2).value = 'IBM'
sheet1.cell(2, 3).value = 11

sheet1.cell(3, 1).value = 'SHIVU'
sheet1.cell(3, 2).value = 'CARELON'
sheet1.cell(3, 3).value = 12

sheet1.cell(4, 1).value = 'NAGU'
sheet1.cell(4, 2).value = 'IBM'
sheet1.cell(4, 3).value = 8

redFill = PatternFill(start_color="ff0000",
                      end_color="ff0000",
                      fill_type="solid")
sheet1.cell(1,1).fill = redFill
workbook2.save(file1)
